from openpyxl import Workbook
from collections import namedtuple
from tkinter.filedialog import askopenfilename
from lxml import etree

__author__ = 'M020240'
"""
This class represents an export of the ECDM elements and relationships

Attributes:

"""


class ECDMSpreadsheet:

    def __init__(self, spreadsheet_name, log_message):
        """
        :param spreadsheet_name: the input spreadsheet
        :param log_message A function reference to the logging widget
        :return: no return
        """
        self.filename = spreadsheet_name
        self.log_message = log_message

        self.element = namedtuple('element', ['Name', 'GUID'])
        self.wb = Workbook()
        self.ws = self.wb.active

        # write the header
        self.ws['A1'] = "GUID"
        self.ws['B1'] = "Name"
        self.ws['C1'] = "Type"
        self.ws['D1'] = "ClientID"
        self.ws['E1'] = "SourceGUID"
        self.ws['F1'] = "SupplierID"
        self.ws['G1'] = "DestGUID"

    def close_spreadsheet(self):
        """
        Close the spreadsheet file.
        :return:
        """
        self.wb.save(self.filename)

    def find_erwin_element(self, item_name):
        return

    def write_ecdm_map(self, entity_map, relationship_map, log_message):
        self.log_message("writing ECDM map")
        row = 2
        rows = str(row)
        for key, value in entity_map.items():
            rows = str(row)
            self.ws['A'+rows] = key
            self.ws['B'+rows] = value[0]
            self.ws['C'+rows] = value[1]
            row += 1
            # log_message("e"+value[0])

        for key, value in relationship_map.items():
            rows = str(row)
            self.ws['A'+rows] = key
            self.ws['B'+rows] = value[0]
            self.ws['C'+rows] = '-'
            self.ws['D'+rows] = value[1]
            self.ws['E'+rows] = value[2]
            self.ws['F'+rows] = value[3]
            self.ws['G'+rows] = value[4]
            row += 1
            # log_message('r'+value[0])
        self.log_message("end writing ECDM map")